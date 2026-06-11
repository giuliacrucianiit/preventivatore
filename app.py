"""
Computo Metrico Web — backend Flask
"""
import json
import io
import os
import glob
from pathlib import Path
from datetime import date

from flask import Flask, jsonify, request, send_file

BASE = Path(__file__).parent
DATA = BASE / "data"
DATA.mkdir(exist_ok=True)

CLIENTI_FILE = DATA / "clienti.json"
IMPOSTAZIONI_FILE = DATA / "impostazioni.json"

IMPOSTAZIONI_DEFAULT = {
    "azienda": {"nome": "", "indirizzo": "", "partita_iva": "", "telefono": "", "email": ""},
}

app = Flask(__name__)
_voci_cache: list[dict] = []

# Svuota cache all'avvio
for f in glob.glob(str(DATA / "*cache*")):
    try: os.remove(f)
    except: pass


# ── LETTURA EXCEL ─────────────────────────────────────────────────

def _cell(row, idx):
    if idx is None or idx >= len(row): return ""
    v = row[idx]
    return str(v).strip() if v is not None else ""

def _parse_prezzo(v) -> float:
    if not v: return 0.0
    try:
        return float(str(v).replace("€","").replace(" ","").replace(".","").replace(",","."))
    except: return 0.0

def _leggi_foglio(ws, nome_sorgente: str, nome_foglio: str) -> list[dict]:
    # Colonne fisse: A=0 codice, B=1 descrizione, C=2 um, D=3 prezzo
    col = {"codice": 0, "descrizione": 1, "um": 2, "prezzo": 3}
    skip_desc = {"descrizione dell'articolo","descrizione","voce","lavorazione"}
    skip_cod  = {"numero d'ordine","codice","cod.","n.","numero"}
    voci = []

    for row in ws.iter_rows(values_only=True):
        if not row or all(v is None for v in row): continue
        codice = _cell(row, col["codice"])
        desc   = _cell(row, col["descrizione"])
        um     = _cell(row, col["um"])
        raw    = row[col["prezzo"]] if col["prezzo"] < len(row) else None
        if raw is None: prezzo = 0.0
        elif isinstance(raw, (int, float)): prezzo = float(raw)
        else: prezzo = _parse_prezzo(str(raw))

        if desc.lower() in skip_desc: continue
        if codice.lower() in skip_cod: continue
        if not desc: continue

        voci.append({
            "codice": codice,
            "descrizione": desc,
            "prezzo": prezzo,
            "um": um,
            "ha_prezzo": prezzo > 0,
            "sorgente": nome_sorgente,
            "foglio": nome_foglio,
        })
    return voci

def carica_prezziario() -> list[dict]:
    global _voci_cache
    if _voci_cache: return _voci_cache
    try:
        import openpyxl
    except ImportError:
        return []
    voci = []
    for percorso in sorted(DATA.glob("*.xlsx")):
        try:
            wb = openpyxl.load_workbook(percorso, read_only=True, data_only=True)
        except: continue
        for nome_foglio in wb.sheetnames:
            voci.extend(_leggi_foglio(wb[nome_foglio], percorso.stem, nome_foglio))
        wb.close()
    _voci_cache = voci
    return voci


# ── CLIENTI / IMPOSTAZIONI ────────────────────────────────────────

def leggi_clienti() -> list:
    if not CLIENTI_FILE.exists(): return []
    return json.loads(CLIENTI_FILE.read_text(encoding="utf-8"))

def scrivi_clienti(c): CLIENTI_FILE.write_text(json.dumps(c, ensure_ascii=False, indent=2), encoding="utf-8")

def leggi_impostazioni() -> dict:
    if not IMPOSTAZIONI_FILE.exists(): return dict(IMPOSTAZIONI_DEFAULT)
    try:
        d = json.loads(IMPOSTAZIONI_FILE.read_text(encoding="utf-8"))
        m = dict(IMPOSTAZIONI_DEFAULT); m.update(d)
        m["azienda"] = {**IMPOSTAZIONI_DEFAULT["azienda"], **d.get("azienda",{})}
        return m
    except: return dict(IMPOSTAZIONI_DEFAULT)

def scrivi_impostazioni(s): IMPOSTAZIONI_FILE.write_text(json.dumps(s, ensure_ascii=False, indent=2), encoding="utf-8")


# ── EXPORT PDF ────────────────────────────────────────────────────

def genera_pdf(doc: dict, impostazioni: dict) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_LEFT
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable

    buf = io.BytesIO()
    az = impostazioni.get("azienda", {})
    righe = doc.get("righe", [])
    cliente = doc.get("cliente", {})
    tipo = doc.get("tipo", "Computo").upper()
    numero = doc.get("numero", "")
    note = doc.get("note", "")
    totale = sum(r["quantita"] * r["prezzo_unitario"] for r in righe)

    def fmt(v): return f"€ {v:,.2f}".replace(",","X").replace(".",",").replace("X",".")
    def p(testo, fs=9, bold=False, color="#333333", align=TA_LEFT):
        return Paragraph(str(testo), ParagraphStyle("_",
            fontSize=fs, fontName="Helvetica-Bold" if bold else "Helvetica",
            textColor=colors.HexColor(color), leading=fs+4, alignment=align))

    pdf = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2.5*cm)
    W = A4[0] - 4*cm
    el = []

    az_lines = [p(az.get("nome",""), 13, True, "#1a1a2e")]
    for k in ("indirizzo","partita_iva","telefono","email"):
        if az.get(k): az_lines.append(p(az[k], 8, color="#666"))
    doc_lines = [p(tipo, 18, True, "#1a1a2e"),
                 p(f"N. {numero}" if numero else "", 9, color="#555"),
                 p(f"Data: {date.today().strftime('%d/%m/%Y')}", 9, color="#555")]

    t = Table([[az_lines, doc_lines]], colWidths=[W*0.55, W*0.45])
    t.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("ALIGN",(1,0),(1,0),"RIGHT")]))
    el += [t, Spacer(1,8), HRFlowable(width="100%",thickness=0.5,color=colors.HexColor("#ccc")), Spacer(1,10)]

    if cliente.get("nome"):
        el.append(p("DESTINATARIO", 7.5, True, "#888"))
        el.append(p(cliente["nome"], 10, True))
        if cliente.get("indirizzo"): el.append(p(cliente["indirizzo"], 9))
        det = [x for x in [f"P.IVA: {cliente.get('partita_iva','')}" if cliente.get("partita_iva") else "",
                            cliente.get("telefono",""), cliente.get("email","")] if x]
        if det: el.append(p("  |  ".join(det), 8, color="#777"))
        el.append(Spacer(1,12))

    cw = [W*0.14, W*0.38, W*0.06, W*0.08, W*0.17, W*0.17]
    dati = [["Codice","Descrizione","U.M.","Qnt.","Prezzo u.","Totale"]]
    for r in righe:
        tot = r["quantita"] * r["prezzo_unitario"]
        dati.append([p(r.get("codice",""),8), p(r.get("descrizione",""),9),
                     p(r.get("um",""),8), p(f"{r['quantita']:g}",9,align=TA_RIGHT),
                     p(fmt(r["prezzo_unitario"]),9,align=TA_RIGHT),
                     p(fmt(tot),9,True,align=TA_RIGHT)])
    tv = Table(dati, colWidths=cw, repeatRows=1)
    stile = [
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1a1a2e")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,0),8.5),
        ("BOTTOMPADDING",(0,0),(-1,0),7),("TOPPADDING",(0,0),(-1,0),7),
        ("FONTSIZE",(0,1),(-1,-1),9),
        ("TOPPADDING",(0,1),(-1,-1),5),("BOTTOMPADDING",(0,1),(-1,-1),5),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("LINEBELOW",(0,0),(-1,-1),0.3,colors.HexColor("#e0e0e0")),
        ("ALIGN",(3,1),(-1,-1),"RIGHT"),("ALIGN",(3,0),(-1,0),"RIGHT"),
    ]
    for i in range(1, len(dati)):
        if i % 2 == 0: stile.append(("BACKGROUND",(0,i),(-1,i),colors.HexColor("#f8f8f8")))
    tv.setStyle(TableStyle(stile))
    el += [tv, Spacer(1,10)]

    tt = Table([["","","","", p("TOTALE",11,True), p(fmt(totale),11,True,align=TA_RIGHT)]], colWidths=cw)
    tt.setStyle(TableStyle([
        ("ALIGN",(4,0),(-1,-1),"RIGHT"),
        ("BACKGROUND",(4,0),(-1,0),colors.HexColor("#f0f0f0")),
        ("LINEABOVE",(4,0),(-1,0),0.8,colors.HexColor("#555")),
        ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
    ]))
    el.append(tt)

    if note:
        el += [Spacer(1,14), HRFlowable(width="100%",thickness=0.3,color=colors.HexColor("#ddd")),
               Spacer(1,6), p("Note",7.5,True,"#888"), p(note,8.5,color="#555")]

    def footer(c, d):
        c.saveState(); c.setFont("Helvetica",7.5); c.setFillColor(colors.HexColor("#aaa"))
        c.drawString(2*cm,1.2*cm, az.get("nome",""))
        c.drawRightString(A4[0]-2*cm,1.2*cm,f"Pagina {d.page}")
        c.restoreState()

    pdf.build(el, onFirstPage=footer, onLaterPages=footer)
    return buf.getvalue()


# ── EXPORT EXCEL ──────────────────────────────────────────────────

def genera_excel(doc: dict, impostazioni: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Computo"

    az = impostazioni.get("azienda", {})
    righe = doc.get("righe", [])
    cliente = doc.get("cliente", {})
    tipo = doc.get("tipo", "Computo")
    numero = doc.get("numero", "")
    note = doc.get("note", "")

    # Stili
    blu_scuro = "1a1a2e"
    blu_ch = "e3f2fd"
    grigio = "f0f0f0"
    bordo = Side(style="thin", color="CCCCCC")
    bordo_cell = Border(left=bordo, right=bordo, top=bordo, bottom=bordo)

    def bold(size=10, color="000000"):
        return Font(bold=True, size=size, color=color)
    def normal(size=10):
        return Font(size=size)

    row = 1

    # Intestazione azienda
    if az.get("nome"):
        ws.cell(row, 1, az["nome"]).font = bold(13)
        row += 1
    for k in ("indirizzo","partita_iva","telefono","email"):
        if az.get(k):
            ws.cell(row, 1, az[k]).font = normal(9)
            row += 1

    row += 1

    # Tipo documento + data
    ws.cell(row, 1, f"{tipo}{' N. ' + numero if numero else ''}").font = bold(14)
    ws.cell(row, 4, f"Data: {date.today().strftime('%d/%m/%Y')}").font = normal(10)
    row += 1; row += 1

    # Cliente
    if cliente.get("nome"):
        ws.cell(row, 1, "DESTINATARIO:").font = bold(9, "888888")
        row += 1
        ws.cell(row, 1, cliente["nome"]).font = bold(11)
        row += 1
        if cliente.get("indirizzo"):
            ws.cell(row, 1, cliente["indirizzo"]).font = normal(9)
            row += 1
        det = []
        if cliente.get("partita_iva"): det.append(f"P.IVA: {cliente['partita_iva']}")
        if cliente.get("telefono"): det.append(cliente["telefono"])
        if cliente.get("email"): det.append(cliente["email"])
        if det:
            ws.cell(row, 1, "  |  ".join(det)).font = normal(9)
            row += 1
        row += 1

    # Intestazioni tabella
    intestazioni = ["Codice", "Descrizione", "U.M.", "Quantità", "Prezzo unitario", "Importo"]
    col_widths   = [18, 60, 8, 12, 18, 18]
    for ci, (h, w) in enumerate(zip(intestazioni, col_widths), 1):
        c = ws.cell(row, ci, h)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=blu_scuro)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bordo_cell
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row].height = 20
    row += 1

    # Righe dati
    totale = 0.0
    for idx, r in enumerate(righe):
        tot = r["quantita"] * r["prezzo_unitario"]
        totale += tot
        fill = PatternFill("solid", fgColor="F8F8F8") if idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        valori = [r.get("codice",""), r.get("descrizione",""), r.get("um",""),
                  r["quantita"], r["prezzo_unitario"], tot]
        allineamenti = ["left","left","center","right","right","right"]
        for ci, (val, al) in enumerate(zip(valori, allineamenti), 1):
            c = ws.cell(row, ci, val)
            c.font = normal(10)
            c.fill = fill
            c.alignment = Alignment(horizontal=al, vertical="top", wrap_text=ci==2)
            c.border = bordo_cell
            # Formato numerico per prezzi
            if ci in (4, 5, 6):
                c.number_format = '#.##0,00 €' if ci > 4 else '#.##0,##'
        ws.row_dimensions[row].height = 30
        row += 1

    # Riga totale
    row += 1
    c_tot_label = ws.cell(row, 5, "TOTALE")
    c_tot_label.font = bold(11)
    c_tot_label.alignment = Alignment(horizontal="right")
    c_tot_val = ws.cell(row, 6, totale)
    c_tot_val.font = bold(12)
    c_tot_val.number_format = '#.##0,00 €'
    c_tot_val.alignment = Alignment(horizontal="right")
    c_tot_val.fill = PatternFill("solid", fgColor=grigio)

    # Note
    if note:
        row += 2
        ws.cell(row, 1, "Note:").font = bold(9, "888888")
        row += 1
        c = ws.cell(row, 1, note)
        c.font = normal(9)
        ws.merge_cells(f"A{row}:F{row}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── ROUTES ────────────────────────────────────────────────────────

@app.route("/")
def index():
    return open(BASE / "index.html", encoding="utf-8").read()

@app.route("/api/prezziario")
def api_prezziario():
    q = request.args.get("q","").lower().strip()
    sorgente = request.args.get("sorgente","")
    voci = carica_prezziario()
    if q:
        tokens = q.split()
        voci = [v for v in voci
                if all(t in f"{v['codice']} {v['descrizione']}".lower() for t in tokens)
                and (not sorgente or v.get("sorgente") == sorgente)]
    elif sorgente:
        voci = [v for v in voci if v.get("sorgente") == sorgente]
    else:
        voci = voci[:300]
    return jsonify(voci[:500])

@app.route("/api/prezziario/sorgenti")
def api_sorgenti():
    voci = carica_prezziario()
    return jsonify(sorted(set(v.get("sorgente","") for v in voci if v.get("sorgente"))))

@app.route("/api/clienti", methods=["GET"])
def api_clienti_get(): return jsonify(leggi_clienti())

@app.route("/api/clienti", methods=["POST"])
def api_clienti_post():
    c = leggi_clienti(); c.append(request.json); scrivi_clienti(c)
    return jsonify({"ok": True, "index": len(c)-1})

@app.route("/api/clienti/<int:idx>", methods=["PUT"])
def api_clienti_put(idx):
    c = leggi_clienti()
    if 0 <= idx < len(c): c[idx] = request.json; scrivi_clienti(c)
    return jsonify({"ok": True})

@app.route("/api/clienti/<int:idx>", methods=["DELETE"])
def api_clienti_delete(idx):
    c = leggi_clienti()
    if 0 <= idx < len(c): del c[idx]; scrivi_clienti(c)
    return jsonify({"ok": True})

@app.route("/api/impostazioni", methods=["GET"])
def api_impostazioni_get(): return jsonify(leggi_impostazioni())

@app.route("/api/impostazioni", methods=["POST"])
def api_impostazioni_post(): scrivi_impostazioni(request.json); return jsonify({"ok": True})

@app.route("/api/pdf", methods=["POST"])
def api_pdf():
    try:
        pdf_bytes = genera_pdf(request.json, leggi_impostazioni())
        return send_file(io.BytesIO(pdf_bytes), mimetype="application/pdf",
                         as_attachment=True,
                         download_name=f"computo_{date.today().strftime('%Y%m%d')}.pdf")
    except Exception as e:
        return jsonify({"errore": str(e)}), 500

@app.route("/api/excel", methods=["POST"])
def api_excel():
    try:
        xlsx_bytes = genera_excel(request.json, leggi_impostazioni())
        return send_file(io.BytesIO(xlsx_bytes),
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True,
                         download_name=f"computo_{date.today().strftime('%Y%m%d')}.xlsx")
    except Exception as e:
        return jsonify({"errore": str(e)}), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"\n✓ Computo avviato → http://localhost:{port}\n")
    app.run(host="0.0.0.0", port=port)
